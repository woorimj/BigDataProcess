#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename="studentTest.xlsx")
sheet = wb.active

data = list(sheet.iter_rows(values_only=True))

# total 점수를 6번째 열에 업데이트
for i, row in enumerate(data):
    if i > 0:  # 헤더 행을 제외하고 계산
        midterm, final, hw, attendance = row[2], row[3], row[4], row[5]
        total_score = (
            (midterm * 0.3 / 100)
            + (final * 0.35 / 100)
            + (hw * 0.34 / 100)
            + (attendance * 0.01)
        )
        total_score *= 100  # 100으로 다시 곱해서 점수 맞추기
        row = list(row)  # 튜플을 리스트로 변환
        row[6] = total_score  # total 점수를 업데이트
        data[i] = tuple(row)  # 리스트를 튜플로 다시 변환

# 엑셀에 데이터 입력
for i in range(2, len(data) + 2):
    row_data = data[i - 2]
    for j in range(1, len(row_data) + 1):
        sheet.cell(row=i, column=j, value=row_data[j - 1])

# 엑셀 파일 저장
wb.save("studentTest.xlsx")
