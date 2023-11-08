#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename="student.xlsx")
sheet = wb.active

data = list(sheet.iter_rows(values_only=True))

students = len(data) - 1  # header를 제외한 총 학생 수
# students = 74  # header를 제외한 총 학생 수

# total 값으로 내림차순 정렬
data = sorted(data[1:], key=lambda row: row[6], reverse=True)

# 등급 비율계산
A = int(students * 0.3)
A_plus = int(students * 0.3 * 0.5)
B = int(students * 0.7)
B_plus = int(students * 0.7 * 0.5)
C = students - A - B
C_plus = min(int(C * 0.5), C)

grades = ["A+", "A0", "B+", "B0", "C+", "C0", "F"]

for i in range(len(data)):
    row = list(data[i])
    if row[6] < 40:
        grade = "F"
    elif i < A:
        if i < A_plus:
            grade = grades[0]
        else:
            grade = grades[1]
    elif i < B:
        if i < A + B_plus:
            grade = grades[2]
        else:
            grade = grades[3]
    elif i < students - C:
        if i < students - C + C_plus:
            grade = grades[4]
        else:
            grade = grades[5]
    else:
        grade = grades[6]
    row[7] = grade
    data[i] = tuple(row)

# 학번으로 다시 오름차순 정렬
# data = sorted(data, key=lambda row: row[0])

# 테스트용으로 등급으로 정렬해봄
data = sorted(data, key=lambda row: grades.index(row[7]))

# 엑셀에 데이터 입력
for i in range(2, len(data) + 2):
    row_data = data[i - 2]
    for j in range(1, len(row_data) + 1):
        sheet.cell(row=i, column=j, value=row_data[j - 1])

# 엑셀 파일 저장
wb.save("student.xlsx")
